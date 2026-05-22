from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Roteiro de Teste"

# Cores
PURPLE      = "6C3FC5"
PURPLE_LIGHT= "F0EAFA"
PURPLE_MID  = "D4BBEF"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "FAFAFA"
GRAY_BORDER = "E0E0E8"
BLACK       = "1A1A2E"
GREEN_BG    = "EAFAF1"
RED_BG      = "FDECEA"
ORANGE_BG   = "FEF3E2"

thin = Side(style="thin", color=GRAY_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── TÍTULO ──────────────────────────────────────────────
ws.merge_cells("A1:G1")
ws["A1"] = "Roteiro de Teste — Clara IA"
ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:G2")
ws["A2"] = "Preencha a coluna 'Resposta da Clara' e marque o Status após cada teste"
ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
ws["A2"].fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws.row_dimensions[2].height = 22

# ── CABEÇALHO ────────────────────────────────────────────
headers = ["#", "Bloco", "Pergunta para digitar", "O que a resposta deve conter", "Resposta da Clara", "Status", "Observações"]
col_widths = [4, 22, 38, 42, 42, 12, 28]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=PURPLE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 28

# ── DADOS ────────────────────────────────────────────────
rows = [
    # (num, bloco, pergunta, esperado)
    ("1",  "Bloco 1 — Crítico",        "O plano é mensal ou pago uma vez no ano?",                              "R$2.159,88/ano, parcelável em 12x de R$179,99 no cartão, não é mensalidade renovável"),
    ("2",  "Bloco 1 — Crítico",        "Esse valor de R$179,99, eu pago todo mês?",                             "Explicar que é parcelamento anual, não boleto mensal"),
    ("3",  "Bloco 1 — Crítico",        "Preciso de CNPJ para usar o sistema?",                                  "Não precisa, funciona como pessoa física"),
    ("4",  "Bloco 1 — Crítico",        "E para integrar com o iFood, precisa de CNPJ?",                         "O iFood é que exige CNPJ/MEI, não o CardápioWeb"),
    ("5",  "Bloco 2 — iFood",          "Vocês têm integração com o iFood?",                                     "Módulo adicional R$29,99/mês, pedidos entram automaticamente"),
    ("6",  "Bloco 2 — iFood",          "Como funciona essa integração com o iFood? É automática?",              "Configuração única, depois pedidos entram automático"),
    ("7",  "Bloco 2 — iFood",          "A Entrega Fácil do iFood funciona com vocês?",                          "Sim — iFood Sob Demanda, faz cotação antes de solicitar"),
    ("8",  "Bloco 2 — iFood",          "E com o 99 Delivery tem integração?",                                   "Sim, precisa do módulo iFood contratado"),
    ("9",  "Bloco 3 — Pagamento",      "O Pix do cliente vai direto pra minha conta?",                          "Passa pela TUNA PAGAMENTOS, repasse em até 1 dia útil, taxa R$0,50"),
    ("10", "Bloco 3 — Pagamento",      "Quais formas de pagamento o cliente pode usar?",                        "Pix automático, cartão online (MP/Cielo), dinheiro, maquininha TEF"),
    ("11", "Bloco 3 — Hardware",       "Precisa de impressora para usar o sistema?",                            "Não é obrigatório; impressão automática exige Windows + QZ Tray"),
    ("12", "Bloco 3 — Hardware",       "Funciona só no celular ou precisa de computador?",                      "Básico no celular; impressão automática e chatbot exigem Windows"),
    ("13", "Bloco 3 — Hardware",       "Funciona no iPad?",                                                     "Sim, roda no navegador (Safari, Chrome), precisa de internet estável"),
    ("14", "Bloco 4 — Outros",         "Tem taxa de instalação?",                                               "Não, só a mensalidade do plano"),
    ("15", "Bloco 4 — Outros",         "Vocês têm aplicativo para baixar?",                                     "Não, é plataforma online via navegador"),
    ("16", "Bloco 4 — Outros",         "O sistema emite nota fiscal?",                                          "Sim, módulo Fiscal R$69,99/mês, emite apenas NFC-e"),
    ("17", "Bloco 4 — Outros",         "Tem integração com balança para selfservice?",                          "Não tem integração direta; dá para cadastrar produto por peso manualmente"),
    ("18", "Bloco 4 — Outros",         "Qual a diferença de vocês para outros sistemas?",                       "Chatbot WhatsApp incluso, iFood centralizado, Pix automático, NFC-e, fidelidade nativo"),
]

# Cores dos blocos
BLOCO_CORES = {
    "Bloco 1 — Crítico":    "FCE8E8",
    "Bloco 2 — iFood":      "E8F0FC",
    "Bloco 3 — Pagamento":  "E8FCF0",
    "Bloco 3 — Hardware":   "E8FCF0",
    "Bloco 4 — Outros":     "FEF9E8",
}

for i, (num, bloco, pergunta, esperado) in enumerate(rows):
    row = i + 4
    bg = BLOCO_CORES.get(bloco, GRAY_LIGHT)
    row_bg = PatternFill("solid", fgColor=bg)
    empty_bg = PatternFill("solid", fgColor="FFFFFF")

    values = [num, bloco, pergunta, esperado, "", "", ""]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, color=BLACK)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        # Colunas de preenchimento ficam brancas
        if col_idx in (5, 6, 7):
            cell.fill = empty_bg
        else:
            cell.fill = row_bg

    ws.row_dimensions[row].height = 52

# ── VALIDAÇÃO DROPDOWN para Status ───────────────────────
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(
    type="list",
    formula1='"✅ OK,❌ Falhou,⚠️ Parcial"',
    allow_blank=True,
    showDropDown=False
)
dv.sqref = f"F4:F{3 + len(rows)}"
ws.add_data_validation(dv)

# ── LEGENDA ──────────────────────────────────────────────
legend_row = 4 + len(rows) + 1
ws.merge_cells(f"A{legend_row}:G{legend_row}")
ws[f"A{legend_row}"] = "Legenda de Status:    ✅ OK = resposta correta e completa     ❌ Falhou = respondeu errado ou 'vou verificar'     ⚠️ Parcial = resposta incompleta"
ws[f"A{legend_row}"].font = Font(name="Calibri", size=9, italic=True, color="666666")
ws[f"A{legend_row}"].fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
ws[f"A{legend_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws.row_dimensions[legend_row].height = 22

# ── FREEZE E ZOOM ────────────────────────────────────────
ws.freeze_panes = "A4"
ws.sheet_view.zoomScale = 90

# Salva
path = r"c:\Users\alexa\Desktop\Clara IA\teste-clara-ia.xlsx"
wb.save(path)
print(f"Planilha salva: {path}")
